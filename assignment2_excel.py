# coding:utf8
# Python3.5
# 第二次作业，将多个Excel表格里面的内容合并到一个文件中。作业发送到 ceo@iqianyue.com

# 使用的是openpyxl库，所以需要事先下载
import openpyxl

# 将三个有数据的excel工作薄(workbook)打开,每个工作薄中有一张工作表
wb1 = openpyxl.load_workbook("kijidata1.xlsx")
wb2 = openpyxl.load_workbook("kijidata2.xlsx")
wb3 = openpyxl.load_workbook("kijidata3.xlsx")

# 创建一个新的工作薄用来保存合并后的数据，避免污染原始工作簿
wb = openpyxl.Workbook()
# active sheet 是一个工作簿被打开时，最上层的那张工作表
wsheet = wb.active

# 得到一个工作表中的列数
def get_column(wb):
    sheet = wb.active
    column = sheet.max_column
    return column

# 将一张现有工作表追加到新的工作表wsheet中
def do_append(wb,column):
    ws = wb.active
    for row_obj in tuple(ws.rows):
        t = list(row_obj)
        temp = []
        for j in range(0,column):
            temp.append(t[j].value)
        wsheet.append(temp)
    return 0

# 得到三个工作表的列数
column1 = get_column(wb1)
column2 = get_column(wb2)
column3 = get_column(wb3)

# 将三张表的内容合并到一张表
do_append(wb1,column1)
do_append(wb2,column2)
do_append(wb3,column3)


# 保存新工作簿
wb.save("finalsheet.xlsx")





